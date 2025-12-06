from flask import *
from flask_mysqldb import MySQL   
from MySQLdb.cursors import DictCursor                               
import os     
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from io import BytesIO

app=Flask(__name__)

# XAMPP MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'       # Default XAMPP username
app.config['MYSQL_PASSWORD'] = ''       # Empty password by default
app.config['MYSQL_DB'] = 'smartbank'    # Replace with your actual database name
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

mysql = MySQL(app)
app.secret_key = 'abcde12345'  

#------------------------------------------------ HOME PAGE----------------------------------
@app.route("/")
def index():
    return render_template("index.html")


#------------------------------------------------ CUSTOMER DETAILS----------------------------------
@app.route("/view_customer")
def view_details():
    if not session.get('loggedin') or session.get('role') != 'customer':
        flash('Please login to access this page', 'warning')
        return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(cursorclass=DictCursor)
    cur.execute('SELECT * FROM customers WHERE email = %s', (session['email'],))
    user = cur.fetchone()
    cur.close()
    
    return render_template("view_customer.html", user=user)


#--------------------------------------------TRANSFER AMOUNT ----------------------------------
@app.route("/trans_amount", methods=["GET", "POST"])
def transfer_amount():
    if not session.get('loggedin') or session.get('role') != 'customer':
        flash('Please login to access this page', 'warning')
        return redirect(url_for('login'))

    if request.method == 'POST':
        to_account = request.form.get('account_number')
        amt = request.form.get('amount')

        # Basic validation
        if not to_account or not amt:
            flash('Please provide account number and amount', 'danger')
            return render_template('transfer_amount.html', balance=0)

        try:
            amount = float(amt)
         
        except ValueError:
            flash('Invalid amount provided', 'danger')
            return render_template('transfer_amount.html', balance=0)

        if amount <= 0:
            flash('Transfer amount must be greater than zero', 'danger')
            return render_template('transfer_amount.html', balance=0)

        cur = mysql.connection.cursor(cursorclass=DictCursor)
        try:
            # Get sender account_number from customers table (using session email)
            cur.execute('SELECT account_number FROM customers WHERE email = %s', (session['email'],))
            cust = cur.fetchone()
            if not cust:
                flash('Sender account not found', 'danger')
                return redirect(url_for('login'))

            sender_account = cust['account_number']
            
            # Fetch balances from accounts table
            cur.execute('SELECT account_number, balance FROM accounts WHERE account_number = %s', (sender_account,))
            sender_acc = cur.fetchone()
            # determine current balance to pass back to template on errors
            balance = 0

            if sender_acc and sender_acc.get('balance') is not None:
                balance = sender_acc['balance']

            cur.execute('SELECT account_number, balance FROM accounts WHERE account_number = %s', (to_account,))
            receiver_acc = cur.fetchone()

            if not receiver_acc:
                flash('Target account does not exist', 'danger')
                return render_template('transfer_amount.html', balance=balance)

            if not sender_acc:
                flash('Sender account not found in accounts table', 'danger')
                return render_template('transfer_amount.html', balance=0)

            if sender_acc['account_number'] == receiver_acc['account_number']:
                flash('Cannot transfer to the same account', 'warning')
                return render_template('transfer_amount.html', balance=balance)

            # Check sufficient funds (handle missing/NULL balance safely)
            if sender_acc.get('balance') is None or float(sender_acc['balance']) < amount:
                flash('Insufficient funds', 'danger')
                return render_template('transfer_amount.html', balance=balance)

            # Perform transfer within a transaction
            try:
                # Debit sender
                cur.execute('UPDATE accounts SET balance = balance - %s WHERE account_number = %s', (amount, sender_acc['account_number']))
                # Credit receiver
                cur.execute('UPDATE accounts SET balance = balance + %s WHERE account_number = %s', (amount, receiver_acc['account_number']))

                # Record transaction (single record covers both sender and receiver)
                cur.execute(
                    'INSERT INTO transactions (from_account, to_account, amount) VALUES (%s, %s, %s)',
                    (sender_acc['account_number'], receiver_acc['account_number'], amount)
                )

                mysql.connection.commit()
                flash('Transfer completed successfully', 'success')
                return redirect(url_for('transactions'))
            except Exception:
                mysql.connection.rollback()
                flash('An error occurred during transfer', 'danger')
                return render_template('transfer_amount.html', balance=balance)

        finally:
            cur.close()

    # GET request: fetch and display sender's current balance
    cur = mysql.connection.cursor(cursorclass=DictCursor)
    try:
        cur.execute('SELECT account_number FROM customers WHERE email = %s', (session['email'],))
        cust = cur.fetchone()
        if not cust:
            # If customer not found, treat balance as zero
            balance = 0
            return render_template('transfer_amount.html', balance=balance)

        cur.execute('SELECT balance FROM accounts WHERE account_number = %s', (cust['account_number'],))
        acc = cur.fetchone()
        if acc and acc.get('balance') is not None:
            balance = acc['balance']
        else:
            # No account row -> treat balance as zero
            balance = 0

        return render_template('transfer_amount.html', balance=balance)
    except Exception as e:
        print('transfer_amount GET error:', repr(e))
        return render_template('transfer_amount.html', balance=0)
    finally:
        cur.close()



#--------------------------------------------TRANSACTIONS ----------------------------------
@app.route("/transactions")
def transactions():
    if not session.get('loggedin') or session.get('role') != 'customer':
        flash('Please login to access this page', 'warning')
        return redirect(url_for('login'))

    cur = mysql.connection.cursor(cursorclass=DictCursor)
    try:
        # Get the logged-in customer's account number
        cur.execute('SELECT account_number FROM customers WHERE email = %s', (session['email'],))
        cust = cur.fetchone()
        if not cust:
            flash('Customer not found', 'danger')
            return redirect(url_for('login'))

        acct = cust['account_number']

        # Query transactions where this account was sender or receiver
        # Assumes a `transactions` table with columns: id, from_account, to_account, amount, timestamp
        cur.execute('''
            SELECT id, from_account, to_account, amount, timestamp
            FROM transactions
            WHERE from_account = %s OR to_account = %s
            ORDER BY timestamp DESC
        ''', (acct, acct))
        txns = cur.fetchall()

        return render_template('transactions.html', transactions=txns)
    except Exception:
        flash('Could not load transactions', 'danger')
        return render_template('transactions.html', transactions=[])
    finally:
        cur.close()


# -------------------------------SAVINGS-----------------------------------------------------
@app.route("/savings", methods=["GET", "POST"])
def savings():
    if not session.get('loggedin') or session.get('role') != 'customer':
        flash('Please login to access this page', 'warning')
        return redirect(url_for('login'))

    if request.method == 'POST':
        amt = request.form.get('savings_amount')
        # Basic validation
        if not amt:
            flash('Please provide amount', 'danger')
            return render_template('savings.html', balance=0)
        try:
            amount = float(amt)
        except ValueError:
            flash('Invalid amount provided', 'danger')
            return render_template('savings.html', balance=0)
        if amount <= 0:
            flash('Amount must be greater than zero', 'danger')
            return render_template('savings.html', balance=0)

        cur = mysql.connection.cursor(cursorclass=DictCursor)
        try:
            # Get customer's account_number from customers table (using session email)
            cur.execute('SELECT account_number FROM customers WHERE email = %s', (session['email'],))
            cust = cur.fetchone()
            if not cust:
                flash('Account not found', 'danger')
                return redirect(url_for('login'))
            account_number = cust['account_number']
            
            # Update balance in accounts table
            cur.execute('UPDATE accounts SET balance = balance + %s WHERE account_number = %s', (amount, account_number))
            
            # Record savings transaction
            cur.execute(
                'INSERT INTO transactions (from_account, to_account, amount) VALUES (%s, %s, %s)',
                (account_number, account_number, amount)
            )
            
            mysql.connection.commit()
            flash('Amount added to savings successfully', 'success')
            
        except Exception:
            mysql.connection.rollback()
            flash('An error occurred while saving', 'danger')
            return render_template('savings.html', balance=0)
        finally:
            cur.close()
    
    # GET request: fetch and display sender's current balance
    cur = mysql.connection.cursor(cursorclass=DictCursor)
    try:
        cur.execute('SELECT account_number FROM customers WHERE email = %s', (session['email'],))
        cust = cur.fetchone()
        # default to zero; if account exists and has a balance use it
        balance = 0
        if cust:
            cur.execute('SELECT balance FROM accounts WHERE account_number = %s', (cust['account_number'],))
            acc = cur.fetchone()
            if acc and acc.get('balance') is not None:
                balance = acc['balance']
        return render_template('savings.html', balance=balance)
    except Exception as e:
        print('savings GET error:', repr(e))
        return render_template('savings.html', balance=0)
    finally:
        cur.close()

#-----------------------------------------CUSTOMER DASHBOARD---------------------------------
@app.route("/customer")
def customer():
    if not session.get('loggedin') or session.get('role') != 'customer':
        flash('Please login to access this page', 'warning')
        return redirect(url_for('login'))
   
    return render_template("customer.html")

#---------------------------------------ADMIN DASHBOARD---------------------------------

@app.route("/admin")
def admin():
    if not session.get('loggedin') or session.get('role') != 'admin':
        flash('Unauthorized access', 'danger')
        return redirect(url_for('login'))
   
    # compute total revenue as sum of all account balances (treat NULL as 0)
    cur = mysql.connection.cursor(cursorclass=DictCursor)
    try:
        cur.execute('SELECT COALESCE(SUM(balance), 0) AS total FROM accounts')
        row = cur.fetchone()
        total = row.get('total') if row else 0
        # format with comma separators and 2 decimals for display
        total_formatted = "{:, .2f}".replace(' ', '') if False else None
        try:
            # prefer numeric formatting with thousands separator
            total_formatted = "{:, .2f}".replace(' ', '')
        except Exception:
            total_formatted = None
        # fallback: if total is numeric, format normally
        if total_formatted is None:
            try:
                total_revenue = "{:, .2f}".replace(' ', '')
            except Exception:
                total_revenue = total
        # simpler: just format using Python's format
        try:
            total_revenue = format(float(total), ',.2f')
        except Exception:
            total_revenue = total
        return render_template("admin.html", total_revenue=total_revenue)
    finally:
        cur.close()

#---------------------------------------CUSTOMER DATABASE---------------------------------
@app.route("/view_cus_db")
def view_cus_db():
    if not session.get('loggedin') or session.get('role') != 'admin':
        flash('Unauthorized access', 'danger')
        return redirect(url_for('login'))
    
    cur2=mysql.connection.cursor()
    cur2.execute("SELECT * FROM customers ORDER BY account_number ASC")
    data2=cur2.fetchall()
    cur2.close()
    return render_template('view_cus_db.html',user_data=data2)


#---------------------------------------EXPORT CUSTOMERS TO EXCEL---------------------------------
@app.route("/export_customers")
def export_customers():
    if not session.get('loggedin') or session.get('role') != 'admin':
        flash('Unauthorized access', 'danger')
        return redirect(url_for('login'))
    
    try:
        # Fetch all customer data
        cur = mysql.connection.cursor(cursorclass=DictCursor)
        cur.execute("SELECT * FROM customers ORDER BY account_number ASC")
        customers = cur.fetchall()
        cur.close()
        
        # Create a new workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        # Add headers
        headers = ['Account Number', 'Username', 'DOB', 'Gender', 'Aadhar', 'PAN', 'Phone', 'Email', 'Address']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for customer in customers:
            row = [
                customer.get('account_number', ''),
                customer.get('username', ''),
                customer.get('dob', ''),
                customer.get('gender', ''),
                customer.get('aadhar_number', ''),
                customer.get('pan_number', ''),
                customer.get('phone_number', ''),
                customer.get('email', ''),
                customer.get('address', '')
            ]
            ws.append(row)
        
        # Adjust column widths
        column_widths = [18, 15, 12, 10, 15, 12, 12, 20, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Save to BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with current date
        filename = f"customers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Return file as download
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print('Export error:', repr(e))
        flash('An error occurred while exporting data', 'danger')
        return redirect(url_for('view_cus_db'))


#-------------------------------------------EXPORT TRANSACTIONS----------------------------------
@app.route("/export_transactions")
def export_transactions():
    if not session.get('loggedin'):
        flash('Please login to access this page', 'warning')
        return redirect(url_for('login'))

    try:
        cur = mysql.connection.cursor(cursorclass=DictCursor)

        # If admin, export ALL transactions; if customer, export only theirs
        if session.get('role') == 'admin':
            cur.execute('SELECT id, from_account, to_account, amount, timestamp FROM transactions ORDER BY timestamp DESC')
            trans = cur.fetchall()
        else:
            # customer
            cur.execute('SELECT account_number FROM customers WHERE email = %s', (session.get('email'),))
            customer = cur.fetchone()
            if not customer:
                flash('Customer account not found', 'danger')
                cur.close()
                return redirect(url_for('transactions'))

            account_number = customer['account_number']
            cur.execute('''
                SELECT id, from_account, to_account, amount, timestamp
                FROM transactions
                WHERE from_account = %s OR to_account = %s
                ORDER BY timestamp DESC
            ''', (account_number, account_number))
            trans = cur.fetchall()

        cur.close()

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ['Date/Time', 'From Account', 'To Account', 'Amount']
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for transaction in trans:
            # Ensure timestamp is string for Excel representation
            ts = transaction.get('timestamp')
            if hasattr(ts, 'strftime'):
                ts = ts.strftime('%Y-%m-%d %H:%M:%S')

            row = [
                ts,
                transaction.get('from_account', ''),
                transaction.get('to_account', ''),
                transaction.get('amount', '')
            ]
            ws.append(row)

        column_widths = [25, 20, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"transactions_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print('Export transactions error:', repr(e))
        flash('An error occurred while exporting transactions', 'danger')
        return redirect(url_for('transactions'))


#-------------------------------------------REGISTERATION PAGE----------------------------------
@app.route('/register', methods=['GET', 'POST'])
def register():
     if (request.method == 'POST'):
        # Get form data
        account_number = request.form.get('account_number')
        username = request.form.get('username')
        dob = request.form.get('dob')
        gender = request.form.get('gender')
        aadhar_number = request.form.get('aadhar_number')
        pan_number = request.form.get('pan_number')
        phone_number = request.form.get('phone_number')
        email = request.form.get('email')
        password = request.form.get('password')
        address = request.form.get('address')

        # Validate required fields
        if not all([account_number, username, dob,gender, aadhar_number, pan_number, phone_number, email, password, address]):
            flash("All fields are required!", "danger")
            return render_template('register.html')

        cur1 = mysql.connection.cursor()
        cur1.execute('SELECT email FROM customers WHERE email = %s OR account_number = %s', (email, account_number))
        existing = cur1.fetchone()

        if existing:
            flash("An account with this email or account number already exists.", "danger")
            return render_template('register.html')
        
        else:
            cur1.execute(
                    "INSERT INTO customers (account_number, username, dob, gender, aadhar_number, pan_number, phone_number, email, password, address) VALUES (%s, %s, %s, %s, %s, %s,%s, %s, %s, %s)",
                    (account_number, username, dob, gender, aadhar_number, pan_number, phone_number, email, password, address)
                )
            cur1.execute(
                    "INSERT INTO accounts (account_number, balance) VALUES (%s, %s)",
                    (account_number, 0)
                )
            mysql.connection.commit()
            cur1.close()
            flash("Registration successful! Please login.", "success")
            return redirect(url_for('login'))
       
     return render_template('register.html') 

        
 #-------------------------------------------LOGIN PAGE----------------------------------   
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        

        # Check for admin login
        if email == "admin@gmail.com" and password == "admin123":
            session['loggedin'] = True
            session['role'] = 'admin'
            session['email'] = email
            session['username'] = 'Admin'
            return redirect(url_for('admin'))
    
        # Check customer login
        cur = mysql.connection.cursor(cursorclass=DictCursor)
        cur.execute("SELECT * FROM customers WHERE email=%s AND password=%s", (email, password))
        data = cur.fetchone()
        cur.close()

        if data:
            session['loggedin'] = True
            session['role'] = 'customer'
            session['username'] = data['username']
            session['email'] = data['email']
            return redirect(url_for('customer'))
        else:
            flash("Invalid Email or Password", "danger")
            return render_template('login.html')
    
    return render_template('login.html')


#---------------------------------------FORGOT PASSWORD---------------------------------
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # Validate inputs
        if not all([email, new_password, confirm_password]):
            flash('Please fill in all fields', 'danger')
            return render_template('forgot_password.html')
            
        if new_password != confirm_password:
            flash('Passwords do not match', 'danger')
            return render_template('forgot_password.html')
            
        # Check if email exists
        cur = mysql.connection.cursor(cursorclass=DictCursor)
        cur.execute('SELECT * FROM customers WHERE email = %s', (email,))
        user = cur.fetchone()
        
        if not user:
            flash('No account found with that email address', 'danger')
            return render_template('forgot_password.html')
            
        try:
            # Update password
            cur.execute('UPDATE customers SET password = %s WHERE email = %s', 
                       (new_password, email))
            mysql.connection.commit()
            flash('Password has been updated successfully! You can now login.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            flash('An error occurred while updating the password', 'danger')
            return render_template('forgot_password.html')
            
        finally:
            cur.close()
            
    return render_template('forgot_password.html')


#------------------------------VIEW CUSTOMER TRANSACTIONS----------------------------------
@app.route('/cus_trans_his')
def cus_trans_his():
    if not session.get('loggedin') or session.get('role') != 'admin':
        flash('Unauthorized access', 'danger')
        return redirect(url_for('login'))
    
    cur3=mysql.connection.cursor()
    cur3.execute("SELECT * FROM transactions")
    trans_his=cur3.fetchall()
    cur3.close()
    return render_template('cus_trans_his.html',trans=trans_his)

#---------------------------------------UPDATE ACCOUNT----------------------------------
@app.route('/update_profile', methods=['GET', 'POST'])
def update_profile():
    if not session.get('loggedin'):
        flash('Please login to update your profile', 'warning')
        return redirect(url_for('login'))
    
    if request.method == 'GET':
        cur = mysql.connection.cursor(cursorclass=DictCursor)
        cur.execute('SELECT * FROM customers WHERE email = %s', (session['email'],))
        user = cur.fetchone()
        cur.close()
        return render_template('update_profile.html', user=user)
    
    if request.method == 'POST':
        username = request.form.get('username')
        dob = request.form.get('dob')
        phone_number = request.form.get('phone_number')
        email = request.form.get('email')
        address = request.form.get('address')
        
        cur = mysql.connection.cursor()
        try:
            # Check if email exists for other users
            if email != session['email']:
                cur.execute('SELECT * FROM customers WHERE email = %s AND email != %s', 
                          (email, session['email']))
                if cur.fetchone():
                    flash('Email already exists', 'danger')
                    return redirect(url_for('update_profile'))
            
            # Update profile details
            cur.execute('''
                UPDATE customers 
                SET username = %s, dob = %s, phone_number = %s, email = %s, address = %s
                WHERE email = %s
            ''', (username, dob, phone_number, email, address, session['email']))
    
            
            mysql.connection.commit()
            session['username'] = username
            session['email'] = email
            
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('customer'))
            
        except Exception as e:
            # Print exception to console for debugging
            print('Update profile error:', repr(e))
            flash('An error occurred while updating your profile', 'danger')
            return redirect(url_for('update_profile'))
            
        finally:
            cur.close()


#--------------------------------------DELETE ACCOUNT----------------------------------
@app.route('/delete_account', methods=['GET', 'POST'])
def delete_account():
    if not session.get('loggedin'):
        flash('Please login to delete your account', 'warning')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        
        # Verify password before deletion
        cur = mysql.connection.cursor(cursorclass=DictCursor)
        try:
            cur.execute('SELECT * FROM customers WHERE email = %s AND password = %s', 
                       (session['email'], password))
            user = cur.fetchone()
            
            if not user:
                flash('Invalid password. Account deletion failed.', 'danger')
                return redirect(url_for('customer'))
            
            # Delete the account
            cur.execute('DELETE FROM customers WHERE email = %s', (session['email'],))
            cur.execute('DELETE FROM accounts WHERE account_number = %s', (user['account_number'],))
            cur.execute('DELETE FROM transactions WHERE from_account = %s OR to_account = %s', (user['account_number'], user['account_number']))
            mysql.connection.commit()
            
            # Clear session and redirect to home
            session.clear()
            flash('Your account has been successfully deleted.', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            flash('An error occurred while deleting your account.', 'danger')
            return redirect(url_for('customer'))
            
        finally:
            cur.close()
    
    return render_template('delete_account.html')


#--------------------------------------LOGOUT----------------------------------
@app.route('/logout')
def logout():
    # Clear all session data
    session.clear()
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('login'))


if __name__=="__main__":
    app.run(debug = True)